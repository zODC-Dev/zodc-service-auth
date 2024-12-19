from io import BytesIO
from fastapi import UploadFile
import openpyxl


class UtilService:
    def __init__(self):
        pass

    async def extract_excel(self, file: UploadFile):
        """Extract excel data in form create"""
        contents = await file.read()
        workbook = openpyxl.load_workbook(
            filename=BytesIO(contents), data_only=True)

        result = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            headers = [cell.value for cell in sheet[1]]

            sheet_result = {header: [] for header in headers}

            for row in sheet.iter_rows(min_row=2):
                for cell, header in zip(row, headers):
                    sheet_result[header].append(
                        str(cell.value) if cell.value is not None else "")

            if not result:
                result = sheet_result
            else:
                for key, value in sheet_result.items():
                    if key in result:
                        result[key].extend(value)
                    else:
                        result[key] = value

                return result
